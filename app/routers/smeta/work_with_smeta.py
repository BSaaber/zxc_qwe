import re

import openpyxl
from openpyxl import load_workbook
from io import BytesIO
from enum import Enum
from app.routers.smeta import schemas
from fastapi import HTTPException, status
import app.database.api as db_api
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Union
from app.database import schemas as db_schemas
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


class ColumnIndexes(BaseModel):
    code: Union[int, None] = None
    name: Union[int, None] = None
    uom: Union[int, None] = None
    amount: Union[int, None] = None
    price: Union[int, None] = None


class ParseMode(Enum):
    FINDING_COLUMNS_INDEXES = 1
    PARSING_LINES = 2


class SmetaLineStandard(Enum):
    TSN = 1
    SN = 2
    UNDEFINED = 3


#              5   .4  -320-7  -1  /1     | 5.4-320-7-1/1
sn_pattern = r'\d+\.\d+-\d+-\d+-\d+/\d+'

#               3  .51 -2  -1             | 3.51-2-1
tsn_pattern = r'\d+.\d+-\d+-\d+'


def get_smeta_standard(code):
    if code is None or not isinstance(code, str):
        return SmetaLineStandard.UNDEFINED
    if re.fullmatch(sn_pattern, code):
        return SmetaLineStandard.SN
    elif re.fullmatch(tsn_pattern, code):
        return SmetaLineStandard.TSN
    else:
        return SmetaLineStandard.UNDEFINED


def cell_is_not_empty_string(cell):
    return isinstance(cell, str) and cell != ''


def cell_is_not_empty_num(cell):
    return isinstance(cell, int) or isinstance(cell, float)


def cell_is_empty(cell):
    return cell is None or cell == '' or (
            not isinstance(cell, str) and not isinstance(cell, int) and not isinstance(cell, float))


def mock(file: bytes):
    result = schemas.Smeta()
    smeta_category = schemas.SmetaCategory()
    smeta_line = schemas.SmetaLine()
    kpgz_piece = db_schemas.KpgzPieceReturn(name="KPGZ PIECE NAME", id=123321)
    spgz_piece = db_schemas.SpgzPieceReturn(name="SPGZ PIECE NAME", id=123321, kpgz_piece=kpgz_piece)
    smeta_hypothesis = db_schemas.HypothesisReturn(spgz_piece=spgz_piece, priority=-1, usage_counter=0)
    for i in range(5):
        smeta_hypothesis.priority = i
        spgz_piece.id = i
        spgz_piece.name = f"name {i}"
        smeta_hypothesis.spgz_piece = spgz_piece.copy()
        print("hypo now:")
        print(smeta_hypothesis)
        smeta_line.hypothesises.append(smeta_hypothesis.copy())
    for i in range(15):
        smeta_line.hypothesises.sort(reverse=True)
        smeta_category.lines.append(smeta_line)
    for i in range(3):
        result.categories.append(smeta_category)

    workbook = load_workbook(BytesIO(file), data_only=True)

    return result


async def patch_smeta(db: Session, path: str, patches: schemas.PatchSmetaIn):
    workbook = load_workbook(path)
    worksheet = workbook.active
    #max_col = worksheet.max_column
    #worksheet.insert_cols(max_col)
    #worksheet.insert_cols(max_col)
    #max_spgz_len = 1
    #max_kpgz_len = 1
    for patch in patches.patches:
        spgz_piece = await db_api.sprav_edit.get_spgz_piece_by_id(db, patch.spgz_id)
        spgz_piece_return = db_schemas.SpgzPieceReturn.from_orm(spgz_piece)
        #if max_spgz_len < len(spgz_piece_return.name):
        #    max_spgz_len = len(spgz_piece_return.name)
        #if max_kpgz_len < len(spgz_piece_return.kpgz_piece.name):
        #    max_kpgz_len = len(spgz_piece_return.kpgz_piece.name)
        worksheet.cell(row=patch.line_number, column=3).value = spgz_piece_return.name
        worksheet.cell(row=patch.line_number, column=4).value = spgz_piece_return.kpgz_piece.name
        worksheet.cell(row=patch.line_number, column=2).value = spgz_piece_return.data_id
    for by_hand_patch in patches.by_hand:
        #if max_spgz_len < len(by_hand_patch.spgz_text):
        #    max_spgz_len = len(by_hand_patch.spgz_text)
        worksheet.cell(row=by_hand_patch.line_number, column=3).value = by_hand_patch.spgz_text
    #worksheet.column_dimensions[get_column_letter(max_col + 1)].width = max_spgz_len
    #worksheet.column_dimensions[get_column_letter(max_col + 2)].width = max_kpgz_len
    key_line_color = PatternFill(start_color='FFD966',
                                 end_color='FFD966',
                                 fill_type='solid')
    for key_line in patches.key_lines:
        for i in range(1, 10):
            worksheet.cell(row=key_line, column=i).fill = key_line_color

    workbook.save(path)
    print("saved")


async def parse_smeta(db: Session, path: str, result_path: str, general_data: schemas.ParseSmetaIn):
    # return mock(file)

    workbook = load_workbook(path, data_only=True)
    #workbook = load_workbook(BytesIO(file), data_only=True)  # file: bytes
    worksheet = workbook.active
    result = schemas.Smeta()
    mode = ParseMode.FINDING_COLUMNS_INDEXES
    searching_address = True
    searching_name = True
    column_indexes = ColumnIndexes()
    lines_with_bad_code_format = []
    a = 0
    b = 0
    total_number_of_lines = 0
    for line_index, row in enumerate(worksheet.iter_rows(values_only=True)):
        if searching_address:
            if line_index >= 30:
                print("address not found")
                searching_address = False
            if cell_is_not_empty_string(row[0]):
                if '???? ????????????' in row[0]:
                    print('found ???? ????????????')
                    result.address = row[0][row[0].find('???? ????????????') + len('???? ????????????'):]
                    if result.address[0] == ':':
                        result.address = result.address[1:]
                    searching_address = False
        if searching_name:
            if line_index >= 30:
                print("name not found")
                searching_name = False
            if cell_is_not_empty_string(row[0]):
                if '????????????????????????' in row[0]:
                    if cell_is_not_empty_string(worksheet.cell(row=line_index, column=1).value):
                        if '??????????' not in worksheet.cell(row=line_index, column=1).value:
                            result.name = worksheet.cell(row=line_index, column=1).value
                            searching_name = False
        if mode is ParseMode.FINDING_COLUMNS_INDEXES:
            print(line_index + 1, "| ", row)
            for j, val in enumerate(row):
                if cell_is_not_empty_string(val):
                    print("val is not empty")
                    if '????????' in val.lower():
                        column_indexes.code = j
                    elif "????????????????????????" in val.lower():
                        column_indexes.name = j
                    elif "????. ??????." == val.lower() or "?????????????? ??????????????????" == val.lower():
                        column_indexes.uom = j
                    elif "??????-????" in val.lower() or "????????????????????" in val.lower():
                        column_indexes.amount = j
                    elif "??????????" in val.lower() and "??????????" in val.lower():
                        column_indexes.price = j
            print("now column_indexes is: ", column_indexes)
            print(column_indexes.dict().values())
            print("---")
            if None not in column_indexes.dict().values():
                mode = ParseMode.PARSING_LINES
                print("MODE CHANGED TO PARSING LINES")
                print("------------------------------")
        elif mode is ParseMode.PARSING_LINES:
            # ?????????????? ????????????????
            if cell_is_not_empty_string(row[0]):
                razdel_word_found = '????????????' in row[0].lower()
                itogo_word_found = '??????????' in row[0].lower()
                smeta_word_found = '????????' in row[0].lower()
                if razdel_word_found and itogo_word_found:
                    if result.categories:
                        print('?? ???????????????? ???????? ??????-???? ????????')
                        if result.categories[-1].total_price == 0:
                            if cell_is_not_empty_num(row[column_indexes.price]):
                                result.categories[-1].total_price = round(row[column_indexes.price], 5)
                            if cell_is_not_empty_num(row[column_indexes.price - 1]):
                                result.categories[-1].total_price = round(row[column_indexes.price - 1], 5)
                                print('???????????????? ???????? ?????? ?????????????? ', result.categories[-1].name)
                elif razdel_word_found and not itogo_word_found:
                    if result.categories:
                        if result.categories[-1].total_price == 0 and not result.categories[-1].lines:
                            print('???????????? ', result.categories[-1].name, '????????, ??????????????')
                            result.categories.pop()
                    print("STARTING NEW CATEGORY")
                    name = row[0].lower()
                    name = name[name.find('????????????') + len('????????????'):]
                    if len(name) != 0 and name[0] == ':':
                        name = name[1:]
                    result.categories.append(schemas.SmetaCategory(name=name))
                    continue
                elif smeta_word_found and itogo_word_found:
                    if cell_is_not_empty_num(row[column_indexes.price]):
                        result.total_price = round(row[column_indexes.price], 5)
                    if cell_is_not_empty_num(row[column_indexes.price - 1]):
                        result.total_price = round(row[column_indexes.price - 1], 5)
            # ?????????????? ????????
            if cell_is_not_empty_string(row[column_indexes.name]):
                if '??????????' in row[column_indexes.name].lower():
                    if result.categories:
                        if result.categories[-1].lines:
                            if cell_is_not_empty_num(row[column_indexes.price]):
                                result.categories[-1].lines[-1].price = round(row[column_indexes.price], 5)
                            if cell_is_not_empty_num(row[column_indexes.price - 1]):
                                result.categories[-1].lines[-1].price = round(row[column_indexes.price - 1], 5)
            # ?????????????? ???????????????? ???????????????????? ????????????
            if cell_is_not_empty_string(row[column_indexes.code]) and get_smeta_standard(
                    row[column_indexes.code] != SmetaLineStandard.UNDEFINED):  # new line
                print("NEW LINE | ", line_index + 1, " | code:", row[column_indexes.code], " | price: ",
                      row[column_indexes.price])
                result.categories[-1].lines.append(schemas.SmetaLine())
                total_number_of_lines += 1
                result.categories[-1].lines[-1].code = row[column_indexes.code]
                result.categories[-1].lines[-1].name = row[column_indexes.name]
                result.categories[-1].lines[-1].uom = row[column_indexes.uom]
                result.categories[-1].lines[-1].amount = row[column_indexes.amount]
                result.categories[-1].lines[-1].code = row[column_indexes.code]
                result.categories[-1].lines[-1].line_number = total_number_of_lines + 15  # line_index + 1

                #if cell_is_not_empty_num(worksheet.cell(row=line_index, column=column_indexes.price - 1).value):
                #    result.categories[-1].lines[-1].line_number = worksheet.cell(row=line_index,
                #                                                                 column=column_indexes.price - 1).value

                standard = get_smeta_standard(result.categories[-1].lines[-1].code)
                print("stadard: ", standard.name)
                print("line: ", result.categories[-1].lines[-1])
                if standard == SmetaLineStandard.SN:
                    sn_piece = await db_api.sprav_edit.get_sn_piece_by_code(db, result.categories[-1].lines[-1].code)
                    if sn_piece is None:
                        print('no such sn')
                        lines_with_bad_code_format.append(line_index + 1)
                        result.categories[-1].lines.pop()
                        total_number_of_lines -= 1
                        continue
                    print("sn piece: ", sn_piece)
                    hypothesises = await db_api.sprav_edit.get_sn_hypothesises_by_sn_id(db, sn_piece.id)
                    if hypothesises is None:
                        print("no hypothesises")
                        # result.categories[-1].lines.pop()
                    else:
                        hypo_dict = {hypo.priority: hypo for hypo in hypothesises}
                        result.categories[-1].lines[-1].hypothesises = list(hypo_dict.values()) # hypothesises
                        result.categories[-1].lines[-1].hypothesises.sort(reverse=True, key=lambda x: x.priority)
                elif standard == SmetaLineStandard.TSN:
                    print("here")
                    tsn_piece = await db_api.sprav_edit.get_tsn_piece_by_code(db, result.categories[-1].lines[-1].code)
                    if tsn_piece is None:
                        print('no such tsn')
                        lines_with_bad_code_format.append(line_index + 1)
                        result.categories[-1].lines.pop()
                        total_number_of_lines -= 1
                        a += 1
                        continue
                    print("tsn piece: ", tsn_piece)
                    hypothesises = await db_api.sprav_edit.get_tsn_hypothesises_by_tsn_id(db, tsn_piece.id)
                    if hypothesises is None:
                        print("no hypothesises")
                        # result.categories[-1].lines.pop()
                    else:
                        hypo_dict = {hypo.priority: hypo for hypo in hypothesises}
                        result.categories[-1].lines[-1].hypothesises = list(hypo_dict.values()) # hypothesises
                        result.categories[-1].lines[-1].hypothesises.sort(reverse=True, key=lambda x: x.priority)
                    b += 1
                elif standard == SmetaLineStandard.UNDEFINED:
                    lines_with_bad_code_format.append(line_index + 1)
            # unimportant line
            else:
                print("TRASH | ", line_index + 1)

    for category in result.categories:
        if category:
            category.key_line = max(category.lines)

    result_workbook = load_workbook(result_path)
    result_worksheet = result_workbook.active
    result_line_number = 16
    for category in result.categories:
        for line in category.lines:
            result_worksheet.cell(row=result_line_number, column=1).value = result_line_number - 15
            result_worksheet.cell(row=result_line_number, column=5).value = line.code
            result_worksheet.cell(row=result_line_number, column=6).value = line.name
            result_worksheet.cell(row=result_line_number, column=7).value = line.uom
            result_worksheet.cell(row=result_line_number, column=8).value = line.amount
            result_worksheet.cell(row=result_line_number, column=9).value = line.price
            result_line_number += 1

    if general_data.name is not None:
        result.name = general_data.name
    if general_data.address is not None:
        result.address = general_data.address

    if result.name is not None:
        result_worksheet.cell(row=3, column=4).value = result.name
    if result.address is not None:
        result_worksheet.cell(row=4, column=4).value = result.address
    if result.total_price is not None:
        result_worksheet.cell(row=5, column=4).value = result.total_price

    result_workbook.save(result_path)

    print(f"\n\n\nTOTAL ERRORS: {len(lines_with_bad_code_format)}\n\n\n")
    print("a", a)
    print("b", b)

    return result
