from openpyxl import load_workbook
from sqlalchemy import create_engine
from app.database import db_models  # noqa - for db initialization
from app.database import schemas as db_schemas
from app.scripts.morphem_handler import TextHandler
import app.database.api as db_api
import os
from dotenv import load_dotenv
import psycopg2  # noqa
from sqlalchemy.orm import sessionmaker
import asyncio

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(BASE_DIR, ".env"))

engine = create_engine(os.environ["DATABASE_URL"])
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


async def parse_spgz_kpgz(filename: str):
    print("parse_spgz_kpgz started to work")
    with SessionLocal() as db:
        text_handler = TextHandler()
        workbook = load_workbook(filename, data_only=True)
        worksheet = workbook.active
        errors = 0
        i = 0
        for row in worksheet.iter_rows(min_row=2, max_col=9, values_only=True):
            i += 1
            if i % 1000 == 0:
                print(i)
            # parse xlsx table
            data_id = row[0]
            kpgz_name = row[3]
            spgz_name = row[4]
            spgz_mapping_info = ','.join(text_handler.process_text(spgz_name))
            spgz_description = row[5]
            uom = row[6]
            okpd = row[7]
            okpd2 = row[8]

            # form kpgz_piece
            kpgz_piece = db_schemas.KpgzPieceCreate(name=kpgz_name)

            # check kpgz_piece in db
            kpgz_result = await db_api.sprav_edit.get_kpgz_piece_by_name(db, kpgz_name)
            if kpgz_result is None:
                # write kpgz_piece to db if needed
                kpgz_result = await db_api.sprav_edit.add_kpgz_piece(db, kpgz_piece)
                if kpgz_result is None:
                    errors += 1
                    print("error during writing kpgz to db")
                    print(row)
                    continue

            # form spgz_piece
            spgz_piece = db_schemas.SpgzPieceCreate(name=spgz_name, mapping_info=spgz_mapping_info, okpd=okpd, okpd2=okpd2, uom=uom,
                                                    description=spgz_description, data_id=data_id,
                                                    kpgz_piece_id=kpgz_result.id)

            # check spgz_piece in db
            spgz_result = await db_api.sprav_edit.get_spgz_piece_by_data_id(db, data_id)
            if spgz_result is None:
                # write spgz_piece to db
                spgz_result = await db_api.sprav_edit.add_spgz_piece(db, spgz_piece)
                if spgz_result is None:
                    errors += 1
                    print("error during writing spgz to db")
                    print(row)
                    continue
        return errors


loop = asyncio.get_event_loop()
errors_amount = loop.run_until_complete(parse_spgz_kpgz("app/scripts/data/spgz_kpgz.xlsx"))
loop.close()
print(errors_amount)
